import os
import shutil
from core.item import Item
from config import constants as cons
from pathlib import Path
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from core.exceptions import (
    ExcelFormatError,                  
    StockShortError,
    FileInuseError,           
    ReadError,  
    WriteError,
)
#==================================================================================
# [1] File path and time: Ensure the program "always works only in its own folder".
# =================================================================================

def now_str() -> str:
    """Current time (string), stored in Last Updated"""  
    DATE_FMT=  cons.DATE_FMT           
    return datetime.now().strftime(DATE_FMT)


def today_ymd() -> str:
    """Today's date (YYYY-MM-DD) will be used in the 'Daily Backup File Name'."""
    return datetime.now().strftime("%Y-%m-%d")

        
class ExcelRepository:
    def __init__(self, path):
        self.path = Path(path)
        self._cache = None

    def complete_the_write(self, mode, row, name, pid, current_qty, qty, receiver, shipper, new_qty) -> Item:
        path = self.file_path()                                   
        self.ensure_filepath(path)                      
        LAST_BACKUP = cons.LAST_BACKUP  
        self.last_backup_path(LAST_BACKUP)              
        self.backup_before_update(path, LAST_BACKUP) 
 
        try:          
            wb, ws = self.load_file_and_list(path)
        except PermissionError:
            raise FileInuseError("⚠ Please close the Excel file while it is in use.")            
        except Exception:                      
            raise ReadError("Failed to read Excel file")  

        if "Log" not in wb.sheetnames:
            log_ws = wb.create_sheet("Log")
            log_ws.append(["Transaction_Time", "Pid", "Name", "Mode", "Quantity", "Buyer"])
            self.sort_log_by_id(ws)
        else:
            log_ws = wb["Log"]
            self.sort_log_by_id(ws)
                      
        write_data = {}              
        delta = qty if mode == "IN" else -qty                 

        if mode == "OUT":                       
            if not receiver: 
                delta = 0
                return        
        if row == ws.max_row + 1:                                                      
            if name is None:
                return                  
            write_data = {
                "Pid": pid,
                "Name": name.strip(),
                "Current_Quantity": delta,
                "Transaction_Time": now_str(),
                "Buyer": receiver,
                "Previous_Quantity": None,
                "Shipper": None
            }
        else:                                                                      
            if new_qty < 0:
                raise StockShortError("Insufficient inventory")           
            write_data = {
                "Pid": pid,
                "Name": name.strip(),
                "Current_Quantity": new_qty,
                "Transaction_Time": now_str(),
                "Buyer": receiver,
                "Previous_Quantity": current_qty,
                "Shipper": shipper
            }                                                                                                             
                 
        # ========= Write to Excel =========
        self.write_item(wb, ws, row, path, write_data)        
        self.sort_sheet_by_id(ws)      
        row = self.find_row_by_id(ws, pid) 
         
        sheet_num = 1
        self.format_sheet(ws, sheet_num)        
            
        if os.path.getsize(path) == 0:
            raise WriteError("Excel write failed (file size is 0)")     
        if not os.path.exists(path):
            return False        
        
        try: 
            self.safe_save_workbook(wb, ws, path)
            self._cache = None
        except Exception:  
            raise FileInuseError("⚠ Please close the Excel file while it is in use.")
                    
        # ========= Write to Log =========
        log_ws = wb["Log"]    
        action = "IN" if delta > 0 else "OUT"
        person = receiver if mode == "OUT" else ""

        # Add a new log entry
        log_ws.append([
            now_str(),
            write_data["Pid"],
            write_data["Name"],
            action,
            abs(delta),
            person
        ])

        # Get the newly added PID    
        pid = write_data["Pid"]               

        #  Sort by "Pid"
        self.sort_log_by_id(log_ws)
        sheet_num = 2
        self.format_sheet(log_ws, sheet_num)
        self.color_all_log_groups(log_ws) 

        # Save the workbook after updating the log
        self.safe_save_workbook(wb, ws, path)
        self._cache = None

        # Return the updated item        
        return Item(             
            pid=pid,
            name=name,
            current_qty=new_qty,
            time_now=now_str(),
            buyer=receiver,
            shipper=shipper,
            row=row
        )   

    def app_dir(self):
        return Path(__file__).resolve().parent.parent / "data"      
    
    def last_backup_path(self, last_backup) -> str:
        """Return the full path to the latest backup file."""
        return os.path.join(self.app_dir(), last_backup)                        
                
    def file_path(self) -> str:
        """Return the full path to the main Excel file."""
        excel_filename = cons.MAIN_FILENAME
        return os.path.join(self.app_dir(), excel_filename) 

    #=======================================================================
    #[2] Excel Initialization/Read: Ensure the file and headers are correct. 
    #=======================================================================
    
    def ensure_filepath(self, path) -> None:      
       
        if os.path.exists(path):
            return
        wb = Workbook()
        ws = wb.active
        ws.title = cons.SHEET_NAME   
        HEADERS = cons.HEADERS 
        ws.append(HEADERS)

        widths = [10, 20, 25, 25, 20, 30, 20]

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(path)

    def safe_save_workbook(self, wb, ws, path: str) -> None:
        """
        Atomic storage (very practical):
        Directly calling wb.save(path) may result in the file being corrupted due to interruption 
        (program shutdown/power failure/antivirus scan).
        Method: First save as a .tmp file, then use os.replace to replace it all at once.
        """        
        tmp = path + ".tmp"
        wb.save(tmp)
        os.replace(tmp, path)

    def load_file_and_list(self, path: str):  
        """
        Open Excel and select "Get Worksheet".
        Also perform "HEADER validation": to prevent someone from manually changing Excel
        columns or writing the program in the wrong place.
        """
        SHEET_NAME = cons.SHEET_NAME 
        HEADERS = cons.HEADERS   

        if self._cache is not None:
            return self._cache  
  
        wb = load_workbook(path)
        # Create the worksheet if it doesn't exist (this usually doesn't happen, 
        # but it's a foolproof precaution).
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
        else:
            ws = wb[SHEET_NAME]

        # Verify that the headers of the first list must be exactly the same.
        first_row = [c.value for c in ws[1]]
        if first_row != HEADERS:
            raise ExcelFormatError(
                "The Excel header format is incorrect.\n"             
                f"Please confirm that column 1 is：{', '.join(HEADERS)}"
            )
        self._cache = (wb, ws)           
        return wb, ws

    
    def backup_before_update(self, path: str, last_backup: str) -> None:    
        # 1. Check if the source file exists
        if not os.path.exists(path):
            print(f"Error: Source file {path} not found.")           
            return        
        # 2. Try copying the file
        try:            
            shutil.copy2(Path(path), Path(path).with_name(last_backup))
            print(f"Success: {path} has been copied to {last_backup}")
        except PermissionError:
            print(
                f"Warning: The file {path} is currently open by another program "
                "(such as Excel) and cannot be copied!"
            )                               
            print("Please close the file and try again...")                           
        return
                        
# ==============================================================
# [3] Excel data manipulation (the concept of Data Access Layer)
# ==============================================================

    def sort_sheet_by_id(self, ws):
        """
        Sort data rows by product ID while keeping the header row unchanged.
        """ 
        data = []
        HEADERS = cons.HEADERS 

        # Read all data rows
        for r in range(2, ws.max_row + 1):
            row_values = []
            empty = True

            for c in range(1, len(HEADERS) + 1):
                val = ws.cell(row=r, column=c).value
                row_values.append(val)
                if val not in (None, ""):
                    empty = False

            if not empty:
                data.append(row_values)

        # Sort by product ID as text, including IDs that start with letters.
        data.sort(key=lambda x: str(x[0]).lower())       
        ws.delete_rows(2, ws.max_row)

        for i, row_data in enumerate(data, start=2):
            for c, val in enumerate(row_data, start=1):
                ws.cell(row=i, column=c).value = val       
        return    

    def format_sheet(self, ws, sheet_num):
        """
        Beautify Excel:
        - Bold header text
        - Center the header
        - Center all columns
        - Freeze the first column
        - Fixed column width
        """       
        HEADERS = cons.HEADERS 

        # Freeze the first column
        ws.freeze_panes = "A2" 

        # Header format
        header_font = Font(size=14, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = center_align
        
        # Center all data
        for r in range(2, ws.max_row + 1):
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=c).alignment = center_align
        # ｓet column width
        if sheet_num == 1:
            widths = [10, 20, 25, 25, 20, 30, 20]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
                
        else:
            if sheet_num == 2:
                widths = [25, 10, 20, 10, 20, 20]
                for i, w in enumerate(widths, start=1):
                    ws.column_dimensions[get_column_letter(i)].width = w
       
    def find_row_by_id(self, ws, pid: str) -> int | None:        
        """
        Find the column containing the ID in the worksheet (start from column 2).
        Returns row index(int) if found, returns None if not found.
        """      
        pid = pid.strip()        
        if not pid:  
            return None

        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value  # column 1 = ID
            if val is None:
                continue
            if str(val).strip() == pid:
                return r           
        return None    
   
    def read_item(self, ws, row: int) -> Item:
        # Read all columns from a given row and return an Item object.               
        pid = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        qty = ws.cell(row=row, column=3).value
        time_now = ws.cell(row=row, column=4).value
        buyer = ws.cell(row=row, column=5).value
        previous_qty = ws.cell(row=row, column=6).value
        shipper = ws.cell(row=row, column=7).value          
        return Item(             
            pid=pid,
            name=name,
            current_qty=qty,
            time_now=time_now,
            buyer=buyer,
            previous_qty=previous_qty,
            shipper=shipper
        )  

    def write_item(self, wb, ws, row: int, path, values: dict) -> None:
        
        HEADERS = cons.HEADERS  
        for i, h in enumerate(HEADERS, start=1):
            ws.cell(row=row, column=i).value = values.get(h)

        self.safe_save_workbook(wb, ws, path)
        return 

    def sort_log_by_id(self, ws):
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        rows.sort(key=lambda r: str(r[1]).lower())   
        ws.delete_rows(2, ws.max_row)

        for r in rows:
            ws.append(r) 

    def color_all_log_groups(self, ws):
        from openpyxl.styles import PatternFill

        current_pid = None
        color_index = -1

        for row in ws.iter_rows(min_row=2):
            pid = row[1].value   

            if pid != current_pid:
                current_pid = pid
                LOG_COLORS = cons.LOG_COLORS  
                color_index = (color_index + 1) % len(LOG_COLORS)
                fill = PatternFill(
                    start_color=LOG_COLORS[color_index],
                    end_color=LOG_COLORS[color_index],
                    fill_type="solid"
                )

            for cell in row:
                cell.fill = fill

    def qty_to_int(self, x) -> int:
        if x is None or str(x).strip() == "":
            return 0
        try:
            return int(x)
        except Exception:
            return 0    